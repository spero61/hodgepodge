# Monthly Calendar Generator for a specific year (.xlsx)
import sys
import datetime
import calendar
import holidays
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side


# create worksheet
def create_month_calendar(month, year):

    kr_holidays = holidays.country_holidays('KR')

    year_str = f"{year}" if year > 0 else f"{abs(year) + 1} BC"

    if month == 1:
        ws = wb.active  # ws: worksheet
        ws.title = f"{calendar.month_abbr[month].upper()}-{year_str}"

    else:
        sheet_name = f"{calendar.month_abbr[month].upper()}-{year_str}"
        ws = wb.create_sheet(sheet_name)

    ws.sheet_properties.tabColor = color_holidays[month]

    # ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    day_list = week_header.split(" ")

    # ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    day_names = list(calendar.day_abbr)

    # 1st day of the month in int: (MONDAY = 0, TUESDAY = 1, ... )
    first_day_idx = calendar.weekday(year=year, month=month, day=1)

    # last day of the month (e.g., 30)
    last_day_int = calendar._monthlen(year=year, month=month)

    # (e.g., Wed)
    first_day_str = day_names[first_day_idx]

    # set named styles: https://openpyxl.readthedocs.io/en/stable/styles.html
    color_holiday = color_holidays[month]
    color_darkgray = "333333"
    font_weekend_head = Font(size=21, bold=True, color=color_holiday)
    font_weekday_head = Font(size=21, bold=True, color=color_darkgray)
    font_holiday = Font(size=27, bold=True, color=color_holiday)
    font_weekday = Font(size=27, bold=True, color=color_darkgray)
    font_month_str = Font(size=58, bold=True, color=color_darkgray)
    font_month_int = Font(size=58, bold=True, color=color_holiday)
    alignment_head = Alignment(horizontal="center", vertical="bottom")
    alignment_day = Alignment(horizontal="center", vertical="top")
    border_top_thick = Border(top=Side(border_style="thick", color=color_darkgray))

    # Month str (e.g., JUNE)
    ws["A1"].value = calendar.month_name[month].upper()
    ws["A1"].font = font_month_str
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Month int (e.g., 6)
    ws["F1"].value = month
    ws["F1"].font = font_month_int
    ws["F1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:E1")
    ws.merge_cells("F1:G1")

    # set row height and column width respectively
    for row in range(ws.max_row):
        if row == 1:  # month
            ws.row_dimensions[row].height = 80.25
        elif row == 2:  # header
            ws.row_dimensions[row].height = 37.5
        else:  # day
            ws.row_dimensions[row].height = 50.25

    column_letters = tuple(
        get_column_letter(col_num + 1) for col_num in range(ws.max_column)
    )
    for column_letter in column_letters:
        ws.column_dimensions[column_letter].width = 11.25

    # day head cells (e.g., Sun, Mon, Tue, ..., Sat)
    day_idx = 0
    for row in ws["A2:G2"]:
        for cell in row:
            cell.value = day_list[day_idx]
            day_idx += 1
            cell.alignment = alignment_head
            if cell.coordinate.startswith("A") or cell.coordinate.startswith("G"):
                cell.font = font_weekend_head
            else:
                cell.font = font_weekday_head

    # match colors for each date cell
    cell_objs = []
    for idx in range(3, 24, 4):
        for row in ws[f"A{idx}:G{idx}"]:
            for cell in row:
                if cell.coordinate.startswith("A"):  # Sundays
                    cell.font = font_holiday
                elif cell.coordinate.startswith("G"):  # Saturdays
                    cell.font = font_holiday
                else:  # Weekdays
                    cell.font = font_weekday
                cell.alignment = alignment_day
                cell_objs.append(cell)

    # write each date to the correspond cell
    # {"Sun": 0, "Mon": 1, "Tue": 2, "Wed": 3, "Thu": 4, "Fri": 5, "Sat": 6}
    day_dict = {}
    for i in range(7):
        day_dict[day_list[i]] = i

    start_day = day_dict[first_day_str]
    current_date = 1
    for i in range(start_day, start_day + last_day_int):
        cell_objs[i].value = current_date
        if datetime.date(year, month, current_date) in kr_holidays:
            cell_objs[i].font = font_holiday
        current_date += 1

    # draw default cell borders for the weekhead row
    for cell in ws["3:3"]:
        cell.border = border_top_thick

    # add thick line as cell borders so as to fit it to the specific calendar
    # e.g., October 2022 requires 6 additional border lines
    #       while February 2026 needs 4 more lines
    year_cell = None
    for idx in range(7, 28, 4):
        if ws[f"A{idx - 4}"].value or ws[f"G{idx - 4}"].value:
            for cell in ws[f"{idx}:{idx}"]:
                cell.border = border_top_thick
                year_cell = cell

    # to add space between rows (by adding a row to the top of row 2)
    ws.insert_rows(2)

    # set print settings
    last_cell = f"{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.print_area = f"A1:{last_cell}"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # year_cell is located below bottom right
    # cf. it might be different from the last_cell
    year_cell.value = year_str
    year_cell.font = Font(size=12, bold="true", color="D7D2CB", italic="true")
    year_cell.alignment = Alignment(horizontal="right", vertical="top")


if __name__ == "__main__":
    dt = datetime.datetime.now()
    active_month = 1

    # Usage: python calendar_gen.py [year] [month]
    if len(sys.argv) == 3:
        try:
            year = int(sys.argv[1])
            month = int(sys.argv[2])
            active_month = month
            if year <= 0:
                print("FYI: Year 0 is 1 BC, year -1 is 2 BC, and so on.")
        except ValueError as err:
            print(err, "You should enter year and month in integer numbers")
            sys.exit(1)
    elif len(sys.argv) == 2:
        try:
            year = int(sys.argv[1])
            month = dt.month
            if year <= 0:
                print("FYI: Year 0 is 1 BC, year -1 is 2 BC, and so on.")
        except ValueError as err:
            print(err, "You should enter year in integer numbers")
            sys.exit(2)
    else:
        year = dt.year
        month = dt.month
        active_month = month

    # color palette for months
    color_holidays = [
        "",
        "B79F7B",  # January
        "9AC1AE",  # February
        "F3A712",  # March
        "E27396",  # April
        "70B77E",  # May
        "6667AB",  # June
        "0074BF",  # July
        "0E958F",  # August
        "FF715B",  # September
        "AA7788",  # October
        "5BC0BE",  # November
        "D90368",  # December
    ]

    # set output_filename (e.g., calendar-2022.xlsx)
    output_filename = f"calendar-{year}.xlsx"

    # to set the first day of the week to Sunday (6)
    calendar.setfirstweekday(calendar.SUNDAY)

    # day_list starts on Sunday.
    week_header = calendar.weekheader(3)

    # create a workbook
    wb = openpyxl.Workbook()  # wb: workbookws

    # create monthly calendar for the year
    for i in range(1, 13):
        create_month_calendar(month=i, year=year)

    # set active sheet other than January if the month argument was given
    ws_idx = active_month - 1
    wb.active = wb.worksheets[ws_idx]

    wb.save(output_filename)
    print(f"[{output_filename}] has created at {dt}")
    wb.close()
